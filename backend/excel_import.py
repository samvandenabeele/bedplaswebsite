import re
from datetime import datetime, date

from extensions import db
from models import Camp, Participant, User


DEFAULT_IMPORTED_USER_PASSWORD = "123456"


def _parse_header_date_range(header):
    date_matches = re.findall(r"\b(\d{2}-\d{2}-\d{4})\b", header)
    if len(date_matches) < 2:
        return None, None

    try:
        start_date = datetime.strptime(date_matches[0], "%d-%m-%Y").date()
        end_date = datetime.strptime(date_matches[1], "%d-%m-%Y").date()
    except ValueError:
        return None, None

    return start_date, end_date


def _extract_camp_metadata(wb):
    header = ""

    for sheet_name in wb.sheetnames:
        value = wb[sheet_name]["A1"].value
        if value:
            header = str(value).strip()
            if header:
                break

    if not header:
        return None

    code_match = re.search(r"\(vakantiecode\s+([^)]+)\)\s*$", header, flags=re.IGNORECASE)
    if code_match is None:
        return None

    code = code_match.group(1).strip()
    name = re.sub(r"\s*\(vakantiecode\s+[^)]+\)\s*$", "", header, flags=re.IGNORECASE).strip() or None
    start_date, end_date = _parse_header_date_range(header)

    camp = Camp.query.filter_by(code=code).first()
    if camp is None:
        camp = Camp(code=code, name=name, source_header=header, start_date=start_date, end_date=end_date)
        db.session.add(camp)
    else:
        if name and not camp.name:
            camp.name = name
        if not camp.source_header:
            camp.source_header = header
        if start_date is not None:
            camp.start_date = start_date
        if end_date is not None:
            camp.end_date = end_date

    db.session.flush()

    return camp


def process_workbook(upload):
    # import lazily so callers can return a helpful error if openpyxl isn't installed
    try:
        from openpyxl import load_workbook
    except Exception:
        raise

    # upload is a FileStorage-like object
    wb = load_workbook(filename=upload, data_only=True)
    camp = _extract_camp_metadata(wb)
    camp_id = camp.id if camp is not None else None

    created_participants = 0
    skipped_participants = 0
    created_counselors = []

    def _cell_str(sheet, row, col):
        v = sheet.cell(row=row, column=col).value
        return str(v).strip() if v is not None else ""

    def _cell_value(sheet, row, col):
        return sheet.cell(row=row, column=col).value

    def _parse_possible_date(v):
        if v in (None, ""):
            return None
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        # Try common string formats
        s = str(v).strip()
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y", "%d %m %Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        # Last resort: try parsing day-month-year with flexible delimiters
        m = re.match(r"^(\d{1,2})[^0-9](\d{1,2})[^0-9](\d{2,4})$", s)
        if m:
            d, mo, y = m.groups()
            if len(y) == 2:
                # assume 20xx for two-digit years
                y = "20" + y
            try:
                return datetime.strptime(f"{d}-{mo}-{y}", "%d-%m-%Y").date()
            except ValueError:
                pass
        return None

    # Process Participants from 'Deelnemers' sheet
    if "Deelnemers" in wb.sheetnames:
        sheet = wb["Deelnemers"]
        row = 11
        while True:
            first = _cell_str(sheet, row, 4)  # D
            last = _cell_str(sheet, row, 5)   # E
            phone_k = _cell_str(sheet, row, 11)  # K
            phone_l = _cell_str(sheet, row, 12)  # L

            if not first and not last:
                # assume end when blank name reached
                break

            if not first or not last:
                skipped_participants += 1
                row += 1
                continue

            phone_1 = phone_k or phone_l
            phone_2 = phone_l if phone_k else ""

            # Try find existing participant by name
            existing = Participant.query.filter(
                Participant.name == first,
                Participant.last_name == last,
            )
            if camp_id is not None:
                existing = existing.filter(Participant.camp_id == camp_id)

            existing = existing.first()

            if existing is None:
                if not phone_1:
                    # require at least one phone to create
                    skipped_participants += 1
                else:
                    birth_val = _cell_value(sheet, row, 2)  # B
                    birth_date = _parse_possible_date(birth_val)

                    new_p = Participant(name=first, last_name=last, phone_1=phone_1, phone_2=phone_2, camp_id=camp_id, birth_date=birth_date)
                    db.session.add(new_p)
                    created_participants += 1
            row += 1

    # Process Counselors from 'Vrijwilligers' sheet
    if "Vrijwilligers" in wb.sheetnames:
        sheet = wb["Vrijwilligers"]
        row = 11

        while True:
            func_val = _cell_str(sheet, row, 2)  # B
            first = _cell_str(sheet, row, 4)     # D
            last = _cell_str(sheet, row, 5)      # E
            email = _cell_str(sheet, row, 13)    # M

            if not first and not last and not func_val and not email:
                break

            # Only add counselors where function is 'Monitor' or contains 'VV'
            func_check = func_val.lower()
            if "monitor" in func_check or "vv" in func_check:
                # Determine username
                base_username = f"{first}.{last}".lower().replace(" ", "")
                username = base_username
                suffix = 1
                while User.query.filter_by(username=username).first() is not None:
                    suffix += 1
                    username = f"{base_username}{suffix}"

                # Only create if email or username not already present
                if User.query.filter_by(email=email).first() is None and User.query.filter_by(username=username).first() is None:
                    u = User(username=username, email=email or None, camp_id=camp_id)
                    u.set_password(DEFAULT_IMPORTED_USER_PASSWORD)
                    u.password_change_required = True
                    db.session.add(u)
                    created_counselors.append({"username": username, "email": email, "password": DEFAULT_IMPORTED_USER_PASSWORD})

            row += 1

    # Commit is left to caller
    return created_participants, skipped_participants, created_counselors

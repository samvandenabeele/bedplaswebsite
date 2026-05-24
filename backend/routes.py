from datetime import datetime, timedelta
from functools import wraps

from flask import Blueprint, current_app, g, jsonify, request
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from sqlalchemy import Date, func, or_

from extensions import db
from models import User, Participant, Water, Urine, Diaper, ClockUse, Clock


api_bp = Blueprint("api", __name__, url_prefix="/api")


def _serializer() -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(current_app.config["SECRET_KEY"], salt="bedplas-auth-token")


def create_access_token(user: User) -> str:
    payload = {"user_id": user.id, "token_version": user.token_version}
    return _serializer().dumps(payload)


def get_user_from_token(token: str) -> User | None:
    try:
        payload = _serializer().loads(token, max_age=current_app.config["AUTH_TOKEN_MAX_AGE"])
    except (BadSignature, SignatureExpired):
        return None

    user = db.session.get(User, payload.get("user_id"))
    if user is None or user.token_version != payload.get("token_version"):
        return None
    return user


def require_auth(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        auth_header = request.headers.get("Authorization", "")
        token = auth_header.removeprefix("Bearer ").strip() if auth_header.startswith("Bearer ") else ""
        if not token:
            return jsonify({"error": "Authorization token required."}), 401

        user = get_user_from_token(token)
        if user is None:
            return jsonify({"error": "Invalid or expired token."}), 401

        g.current_user = user
        return view_func(*args, **kwargs)

    return wrapper


@api_bp.get("/health")
def health_check():
    return jsonify({"status": "ok"})


@api_bp.post("/auth/register")
def register():
    payload = request.get_json(silent=True) or {}
    username = str(payload.get("username", "")).strip().lower()
    email = str(payload.get("email", "")).strip().lower() or None
    password = str(payload.get("password", ""))

    if not username or not password:
        return jsonify({"error": "username and password are required."}), 400

    if User.query.filter_by(username=username).first() is not None:
        return jsonify({"error": "Username already exists."}), 409

    if email and User.query.filter_by(email=email).first() is not None:
        return jsonify({"error": "Email already exists."}), 409

    user = User(username=username, email=email)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()

    token = create_access_token(user)
    return jsonify({"user": user.to_dict(), "token": token}), 201


@api_bp.post("/auth/login")
def login():
    payload = request.get_json(silent=True) or {}
    identifier = str(payload.get("identifier", payload.get("username", payload.get("email", "")))).strip().lower()
    password = str(payload.get("password", ""))

    if not identifier or not password:
        return jsonify({"error": "identifier and password are required."}), 400

    user = User.query.filter(or_(User.username == identifier, User.email == identifier)).first()
    if user is None or not user.check_password(password):
        return jsonify({"error": "Invalid credentials."}), 401

    return jsonify({"user": user.to_dict(), "token": create_access_token(user)})


@api_bp.post("/auth/logout")
@require_auth
def logout():
    user = g.current_user
    user.token_version += 1
    db.session.commit()
    return jsonify({"message": "Logged out successfully."})


@api_bp.get("/auth/me")
@require_auth
def me():
    return jsonify({"user": g.current_user.to_dict()})

@api_bp.route("/addParticipant", methods=["POST"])
@require_auth
def add_participants():
    payload = request.get_json(silent=True)
    if payload is None:
        return jsonify({"error": "incorrect participant info"}), 400

    name = str(payload.get("name", "")).strip()
    if not name:
        return jsonify({"error": "name is required."}), 400
    
    last_name = str(payload.get("last_name", "")).strip()
    if not last_name:
        return jsonify({"error": "last name is required."}), 400
    
    phone_1 = str(payload.get("phone_1", "")).strip()
    if not phone_1:
        return jsonify({"error": "phone_1 is required."}), 400
    
    phone_2 = str(payload.get("phone_2", "")).strip()

    diaper = payload.get("empty_diaper", 0)

    new_participant = Participant(name=name, last_name=last_name, phone_1=phone_1, phone_2=phone_2, empty_diaper=diaper)
    db.session.add(new_participant)
    db.session.commit()
    

    return jsonify({"message": "Participant added successfully.", "name": name}), 201

@api_bp.route("/delParticipant", methods=["POST"])
@require_auth
def del_participant():
    payload = request.get_json(silent=True) or {}
    name = str(payload.get("name", "")).strip()
    last_name = str(payload.get("last_name", "")).strip()

    if not name or not last_name:
        return jsonify({"error": "name and last name are required."}), 400

    participant = Participant.query.filter(
        Participant.name == name,
        Participant.last_name == last_name,
    ).first()

    if participant is None:
        return jsonify({"error": "Participant not found."}), 404

    db.session.delete(participant)
    db.session.commit()

    return jsonify({"message": "Participant deleted successfully."})

@api_bp.route("/queryParticipant", methods=["POST"])
@require_auth
def query_participant():
    payload = request.get_json(silent=True) or {}
    query = Participant.query.filter(Participant.active.is_(True))


    if payload.get("name"):
        query = query.filter(Participant.name.ilike(f"%{payload.get('name')}%"))
    
    if payload.get("last_name"):
        query = query.filter(Participant.last_name.ilike(f"%{payload.get('last_name')}%"))
    
    if payload.get("phone_1"):
        query = query.filter(Participant.phone_1.ilike(f"%{payload.get('phone_1')}%"))
    
    if payload.get("phone_2"):
        query = query.filter(Participant.phone_2.ilike(f"%{payload.get('phone_2')}%"))
    
    participants = query.all()
    today = func.current_date()
    now = datetime.now()
    cutoff_date = now.date() if now.hour >= 18 else now.date() - timedelta(days=1)
    six_pm_today = datetime.combine(cutoff_date, datetime.min.time()).replace(hour=18)
    
    return jsonify({
        "participants": [
            {
                "id": p.id,
                "name": p.name,
                "last_name": p.last_name,
                "phone_1": p.phone_1,
                "phone_2": p.phone_2,
                "empty_diaper": p.empty_diaper,
                "drank_today": db.session.query(func.count(Water.id))
                .filter(
                    Water.participant_id == p.id,
                    func.date(Water.created_at) == today,
                )
                .scalar(),
                "peed_today": db.session.query(func.coalesce(func.sum(Urine.amount), 0))
                .filter(
                    Urine.participant_id == p.id,
                    func.date(Urine.created_at) == today,
                )
                .scalar(),
                "largest_pee": db.session.query(func.coalesce(func.max(Urine.amount), 0))
                .filter(Urine.participant_id == p.id)
                .scalar(),
                "active": p.active,
                "clock": db.session.query(func.count(ClockUse.id))
                .filter(
                    ClockUse.participant_id == p.id,
                    ClockUse.created_at >= six_pm_today,
                )
                .scalar() % 2 == 1,
            }
            for p in participants
        ]
    })

@api_bp.route("/queryCounselor", methods=["POST"])
@require_auth
def query_counselor():
    payload = request.get_json(silent=True) or {}
    # Query users (counselors). There is no `active` on User, so just query User
    query = User.query

    # Accept either `username` or `name` as a search term for the username
    username_term = payload.get("username") or payload.get("name")
    if username_term:
        query = query.filter(User.username.ilike(f"%{username_term}%"))

    if payload.get("email"):
        query = query.filter(User.email.ilike(f"%{payload.get('email')}%"))

    counselors = query.all()

    return jsonify({
        "counselors": [
            {"id": c.id, "username": c.username, "email": c.email, "active": c.active}
            for c in counselors
        ]
    })


@api_bp.patch("/updateEmptyDiaper")
@require_auth
def update_empty_diaper():
    payload = request.get_json(silent=True) or {}
    participant = _resolve_participant(payload)

    if participant is None:
        return jsonify({"error": "Participant not found."}), 404

    try:
        empty_diaper = int(payload.get("empty_diaper"))
    except (TypeError, ValueError):
        return jsonify({"error": "empty_diaper must be an integer."}), 400

    if empty_diaper < 0:
        return jsonify({"error": "empty_diaper must be zero or greater."}), 400

    participant.empty_diaper = empty_diaper
    db.session.commit()

    return jsonify(
        {
            "message": "Empty diaper weight updated successfully.",
            "empty_diaper": participant.empty_diaper,
        }
    )


def _resolve_participant(payload: dict):
    participant_id = payload.get("participant_id")
    if participant_id is not None and str(participant_id).strip() != "":
        try:
            participant = db.session.get(Participant, int(participant_id))
        except (TypeError, ValueError):
            participant = None
        if participant is not None:
            return participant

    name = str(payload.get("name", "")).strip()
    last_name = str(payload.get("last_name", "")).strip()

    if not name or not last_name:
        return None

    return Participant.query.filter(
        Participant.name == name,
        Participant.last_name == last_name,
    ).first()

@api_bp.route("/addWater", methods=["POST"])
@require_auth
def add_water():
    payload = request.get_json(silent=True) or {}
    meal = bool(payload.get("meal"))

    participant = _resolve_participant(payload)

    if participant is None:
        return jsonify({"error": "Participant not found."}), 404

    new_water = Water(participant_id=participant.id, meal=meal)
    db.session.add(new_water)
    db.session.commit()

    return jsonify({"message": "Water entry added successfully."}), 201
    
@api_bp.route("/addUrine", methods=["POST"])
@require_auth
def add_urine():
    payload = request.get_json(silent=True) or {}
    amount = int(payload.get("amount"))
    note = str(payload.get("note", "")).strip() or None

    participant = _resolve_participant(payload)

    if participant is None:
        return jsonify({"error": "Participant not found."}), 404

    new_urine = Urine(participant_id=participant.id, amount=amount, note=note)
    db.session.add(new_urine)
    db.session.commit()

    return jsonify({"message": "Urine entry added successfully."}), 201

@api_bp.route("/addDiaper", methods=["POST"])
@require_auth
def add_diaper():
    payload = request.get_json(silent=True) or {}
    weight = int(payload.get("weight"))
    note = str(payload.get("note", "")).strip() or None

    participant = _resolve_participant(payload)

    if participant is None:
        return jsonify({"error": "Participant not found."}), 404

    liquid_weight = weight - participant.empty_diaper

    new_diaper = Diaper(participant_id=participant.id, weight=liquid_weight, note=note)
    db.session.add(new_diaper)
    db.session.commit()

    return jsonify({"message": "Diaper entry added successfully."}), 201


@api_bp.get("/participantRecentEntries/<int:participant_id>")
@require_auth
def participant_recent_entries(participant_id: int):
    participant = db.session.get(Participant, participant_id)
    if participant is None:
        return jsonify({"error": "Participant not found."}), 404

    try:
        limit = int(request.args.get("limit", 50))
    except (TypeError, ValueError):
        return jsonify({"error": "limit must be an integer."}), 400

    if limit < 1:
        return jsonify({"error": "limit must be greater than 0."}), 400

    # Keep this bounded to protect the endpoint from very large requests.
    limit = min(limit, 200)

    water_entries = (
        Water.query.filter(Water.participant_id == participant_id)
        .order_by(Water.created_at.desc())
        .limit(limit)
        .all()
    )
    urine_entries = (
        Urine.query.filter(Urine.participant_id == participant_id)
        .order_by(Urine.created_at.desc())
        .limit(limit)
        .all()
    )
    diaper_entries = (
        Diaper.query.filter(Diaper.participant_id == participant_id)
        .order_by(Diaper.created_at.desc())
        .limit(limit)
        .all()
    )
    clock_entries = (
        Clock.query.filter(Clock.participant_id == participant_id)
        .order_by(Clock.created_at.desc())
        .limit(limit)
        .all()
    )

    merged_entries = [
        {
            "id": entry.id,
            "kind": "water",
            "created_at": entry.created_at.isoformat() if entry.created_at else None,
            "meal": entry.meal,
            "amount": None,
            "weight": None,
            "note": None,
        }
        for entry in water_entries
    ] + [
        {
            "id": entry.id,
            "kind": "urine",
            "created_at": entry.created_at.isoformat() if entry.created_at else None,
            "meal": None,
            "amount": entry.amount,
            "weight": None,
            "note": entry.note,
        }
        for entry in urine_entries
    ] + [
        {
            "id": entry.id,
            "kind": "diaper",
            "created_at": entry.created_at.isoformat() if entry.created_at else None,
            "meal": None,
            "amount": None,
            "weight": entry.weight,
            "note": entry.note,
        }
        for entry in diaper_entries
    ] + [
        {
            "id": entry.id,
            "kind": "clock",
            "created_at": entry.created_at.isoformat() if entry.created_at else None,
        }
        for entry in clock_entries
    ]

    merged_entries.sort(key=lambda entry: entry.get("created_at") or "", reverse=True)

    return jsonify({"entries": merged_entries[:limit]})


@api_bp.get("/recentEntries")
@require_auth
def recent_entries():
    try:
        limit = int(request.args.get("limit", 100))
    except (TypeError, ValueError):
        return jsonify({"error": "limit must be an integer."}), 400

    if limit < 1:
        return jsonify({"error": "limit must be greater than 0."}), 400

    limit = min(limit, 300)

    water_entries = Water.query.order_by(Water.created_at.desc()).limit(limit).all()
    urine_entries = Urine.query.order_by(Urine.created_at.desc()).limit(limit).all()
    diaper_entries = Diaper.query.order_by(Diaper.created_at.desc()).limit(limit).all()

    participant_ids = {
        entry.participant_id for entry in water_entries + urine_entries + diaper_entries
    }
    participants = Participant.query.filter(Participant.id.in_(participant_ids)).all() if participant_ids else []
    participant_by_id = {
        participant.id: participant
        for participant in participants
    }

    merged_entries = [
        {
            "id": entry.id,
            "kind": "water",
            "participant_id": entry.participant_id,
            "participant_name": participant_by_id.get(entry.participant_id).name
            if participant_by_id.get(entry.participant_id)
            else "Unknown",
            "participant_last_name": participant_by_id.get(entry.participant_id).last_name
            if participant_by_id.get(entry.participant_id)
            else "",
            "created_at": entry.created_at.isoformat() if entry.created_at else None,
            "meal": entry.meal,
            "amount": None,
            "weight": None,
            "note": None,
        }
        for entry in water_entries
    ] + [
        {
            "id": entry.id,
            "kind": "urine",
            "participant_id": entry.participant_id,
            "participant_name": participant_by_id.get(entry.participant_id).name
            if participant_by_id.get(entry.participant_id)
            else "Unknown",
            "participant_last_name": participant_by_id.get(entry.participant_id).last_name
            if participant_by_id.get(entry.participant_id)
            else "",
            "created_at": entry.created_at.isoformat() if entry.created_at else None,
            "meal": None,
            "amount": entry.amount,
            "weight": None,
            "note": entry.note,
        }
        for entry in urine_entries
    ] + [
        {
            "id": entry.id,
            "kind": "diaper",
            "participant_id": entry.participant_id,
            "participant_name": participant_by_id.get(entry.participant_id).name
            if participant_by_id.get(entry.participant_id)
            else "Unknown",
            "participant_last_name": participant_by_id.get(entry.participant_id).last_name
            if participant_by_id.get(entry.participant_id)
            else "",
            "created_at": entry.created_at.isoformat() if entry.created_at else None,
            "meal": None,
            "amount": None,
            "weight": entry.weight,
            "note": entry.note,
        }
        for entry in diaper_entries
    ]

    merged_entries.sort(key=lambda entry: entry.get("created_at") or "", reverse=True)

    return jsonify({"entries": merged_entries[:limit]})


def _entry_model_for_kind(kind: str):
    if kind == "water":
        return Water
    if kind == "urine":
        return Urine
    if kind == "diaper":
        return Diaper
    return None


@api_bp.patch("/entry/<string:kind>/<int:entry_id>")
@require_auth
def update_entry(kind: str, entry_id: int):
    model = _entry_model_for_kind(kind)
    if model is None:
        return jsonify({"error": "Unsupported entry kind."}), 400

    entry = db.session.get(model, entry_id)
    if entry is None:
        return jsonify({"error": "Entry not found."}), 404

    payload = request.get_json(silent=True) or {}

    if kind == "water":
        if "meal" in payload:
            entry.meal = bool(payload.get("meal"))

    if kind == "urine":
        if "amount" in payload:
            try:
                amount = int(payload.get("amount"))
            except (TypeError, ValueError):
                return jsonify({"error": "amount must be an integer."}), 400
            if amount < 0:
                return jsonify({"error": "amount must be zero or greater."}), 400
            entry.amount = amount

        if "note" in payload:
            note = str(payload.get("note", "")).strip() or None
            entry.note = note

    if kind == "diaper":
        if "weight" in payload:
            try:
                weight = int(payload.get("weight"))
            except (TypeError, ValueError):
                return jsonify({"error": "weight must be an integer."}), 400
            if weight < 0:
                return jsonify({"error": "weight must be zero or greater."}), 400
            entry.weight = weight

        if "note" in payload:
            note = str(payload.get("note", "")).strip() or None
            entry.note = note

    db.session.commit()
    return jsonify({"message": "Entry updated successfully."})


@api_bp.delete("/entry/<string:kind>/<int:entry_id>")
@require_auth
def delete_entry(kind: str, entry_id: int):
    model = _entry_model_for_kind(kind)
    if model is None:
        return jsonify({"error": "Unsupported entry kind."}), 400

    entry = db.session.get(model, entry_id)
    if entry is None:
        return jsonify({"error": "Entry not found."}), 404

    db.session.delete(entry)
    db.session.commit()
    return jsonify({"message": "Entry deleted successfully."})

@api_bp.post("/excelParticipantsCounselors")
@require_auth
def excel_participants_counselors():
    # Accept an uploaded Excel file (multipart/form-data, field name 'file') and
    # create Participants and Counselor Users automatically.
    upload = request.files.get("file")
    if upload is None:
        return jsonify({"error": "No file uploaded. Provide file field 'file'."}), 400

    try:
        from openpyxl import load_workbook
    except Exception:
        return jsonify({"error": "openpyxl is not installed on the server."}), 500

    try:
        wb = load_workbook(filename=upload, data_only=True)
    except Exception as exc:
        return jsonify({"error": f"Failed to read workbook: {exc}"}), 400

    created_participants = 0
    skipped_participants = 0
    created_counselors = []

    def _cell_str(sheet, row, col):
        v = sheet.cell(row=row, column=col).value
        return str(v).strip() if v is not None else ""

    # Process Participants from 'Deelnemers' sheet
    if "Deelnemers" in wb.sheetnames:
        sheet = wb["Deelnemers"]
        row = 11
        while True:
            first = _cell_str(sheet, row, 4)  # D
            last = _cell_str(sheet, row, 5)   # E
            phone_k = _cell_str(sheet, row, 11)  # K
            phone_l = _cell_str(sheet, row, 12)  # L

            if not first and not last:
                # assume end when blank name reached
                break

            if not first or not last:
                skipped_participants += 1
                row += 1
                continue

            phone_1 = phone_k or phone_l
            phone_2 = phone_l if phone_k else ""

            # Try find existing participant by name
            existing = Participant.query.filter(
                Participant.name == first,
                Participant.last_name == last,
            ).first()

            if existing is None:
                if not phone_1:
                    # require at least one phone to create
                    skipped_participants += 1
                else:
                    new_p = Participant(name=first, last_name=last, phone_1=phone_1, phone_2=phone_2)
                    db.session.add(new_p)
                    created_participants += 1
            row += 1

    # Process Counselors from 'Vrijwilligers' sheet
    if "Vrijwilligers" in wb.sheetnames:
        sheet = wb["Vrijwilligers"]
        row = 11
        import secrets

        while True:
            func_val = _cell_str(sheet, row, 2)  # B
            first = _cell_str(sheet, row, 4)     # D
            last = _cell_str(sheet, row, 5)      # E
            email = _cell_str(sheet, row, 13)    # M

            if not first and not last and not func_val and not email:
                break

            # Only add counselors where function is 'Monitor' or contains 'VV'
            func_check = func_val.lower()
            if "monitor" in func_check or "vv" in func_check:
                # Determine username
                base_username = f"{first}.{last}".lower().replace(" ", "")
                username = base_username
                suffix = 1
                while User.query.filter_by(username=username).first() is not None:
                    suffix += 1
                    username = f"{base_username}{suffix}"

                # Only create if email or username not already present
                if User.query.filter_by(email=email).first() is None and User.query.filter_by(username=username).first() is None:
                    pwd = secrets.token_urlsafe(10)
                    u = User(username=username, email=email or None)
                    u.set_password(pwd)
                    db.session.add(u)
                    created_counselors.append({"username": username, "email": email, "password": pwd})

            row += 1

    # Commit all new records
    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Database commit failed: {exc}"}), 500

    return jsonify({
        "participants_created": created_participants,
        "participants_skipped": skipped_participants,
        "counselors_created": created_counselors,
    })

@api_bp.post("addClock")
@require_auth
def add_clock():
    payload = request.get_json(silent=True) or {}
    id = int(payload.get("participant_id"))

    new_clock = Clock(participant_id=id)
    db.session.add(new_clock)
    db.session.commit()

    return jsonify({'message' : 'clock added', 'participant_id': id})

@api_bp.post("addClockUse")
@require_auth
def add_clock_use():
    payload = request.get_json(silent=True) or {}
    id = int(payload.get("participant_id"))

    new_clock_use = ClockUse(participant_id=id)
    db.session.add(new_clock_use)
    db.session.commit()

    return jsonify({'message' : 'clockUse added', 'participant_id': id})

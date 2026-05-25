import re
from datetime import datetime, timedelta, time
from pathlib import Path
from shutil import copy2

from sqlalchemy import func
from openpyxl import load_workbook

from extensions import db
from models import Camp, User, Participant, Water, Urine, Diaper, Clock, ClockUse


def _sanitize_filename_part(value):
	value = re.sub(r"[^A-Za-z0-9._ -]", "", str(value or "")).strip()
	value = re.sub(r"\s+", "_", value)
	return value or "participant"


def _as_excel_number(value):
	if isinstance(value, (int, float)):
		return value
	if isinstance(value, str):
		text = value.strip()
		if text == "":
			return ""
		try:
			return int(text)
		except ValueError:
			try:
				return float(text)
			except ValueError:
				return value
	return value


def create_diary(participant):
	base_dir = Path(__file__).resolve().parent
	template_path = base_dir / "static" / "plasdagboek.xlsx"

	participant_name = " ".join(
		part for part in [getattr(participant, "name", ""), getattr(participant, "last_name", "")] if part
	).strip() or f"participant_{getattr(participant, 'id', 'unknown')}"
	safe_participant_name = _sanitize_filename_part(participant_name)
	diary_path = base_dir / "static" / f"plasdagboek_{safe_participant_name}.xlsx"

	copy2(template_path, diary_path)
	wb = load_workbook(filename=diary_path)

	starting_date = getattr(getattr(participant, "camp", None), "start_date", None)
	ending_date = getattr(getattr(participant, "camp", None), "end_date", None)
	clock_use_count = 0

	n_days = 0
	if starting_date is not None and ending_date is not None:
		n_days = max((ending_date - starting_date).days, 0)

	sheet = wb.worksheets["Algemene info"]

	sheet["B3"] = participant.name
	sheet["B4"] = participant.last_name
	sheet["B5"] = participant.birth_date.isoformat() if getattr(participant, "birth_date", None) is not None else ""
	

	for i in range(n_days):
		sheet = wb.worksheets[f"Dag {i}"] if i < len(wb.worksheets) else None

		if starting_date is not None:
			six_pm = datetime.combine(starting_date, time(18, 0))
			six_am_next_day = six_pm + timedelta(hours=12)

			clock_use_count = db.session.query(func.count(ClockUse.id)).filter(
				ClockUse.participant_id == participant.id,
				ClockUse.created_at >= six_pm,
				ClockUse.created_at < six_am_next_day,
			).scalar() or 0

			n_clocks = db.session.query(func.count(Clock.id)).filter(
				Clock.participant_id == participant.id,
				Clock.created_at >= six_pm,
				Clock.created_at < six_am_next_day,
			).scalar() or 0
		
			clock_used = clock_use_count % 2 == 1

			if clock_used:
				sheet["E4"] = "Ja"
			else:
				sheet["E4"] = "Nee"

			sheet["E5"] = n_clocks

			seven_thirty_am_next_day = six_pm + timedelta(hours=13, minutes=30)
			ten_pm_next_day = six_pm + timedelta(days=1, hours=4)
			nine_pm = six_pm + timedelta(hours=3)
			ten_am_next_day = six_pm + timedelta(days=1, hours=4)

			night_urines = db.session.query(Urine).filter(
				Urine.participant_id == participant.id,
				Urine.created_at >= nine_pm,
				Urine.created_at < seven_thirty_am_next_day,
			).order_by(Urine.created_at).all()

			urines = db.session.query(Urine).filter(
				Urine.participant_id == participant.id,
				Urine.created_at >= seven_thirty_am_next_day,
				Urine.created_at < ten_pm_next_day,
			).order_by(Urine.created_at).all()

			waters = db.session.query(Water).filter(
				Water.participant_id == participant.id,
				Water.created_at >= seven_thirty_am_next_day,
				Water.created_at < ten_pm_next_day,
			).order_by(Water.created_at).all()

			diapers = db.session.query(Diaper).filter(
				Diaper.participant_id == participant.id,
				Diaper.created_at >= six_pm,
				Diaper.created_at < ten_am_next_day,
			).order_by(Diaper.created_at).all()

			sheet["E6"] = len(night_urines)
			   
			night_urines_not_op_toilet = [
				u for u in night_urines if (getattr(u, "note", None) or "").strip() != "Op toilet"
			]

			if not diapers and not night_urines_not_op_toilet:
				sheet["E7"] = "Nee"
			else:
				sheet["E7"] = "Ja"

			if diapers:
				sheet["E8"] = _as_excel_number(diapers[-1].empty_weight)
				sheet["E9"] = _as_excel_number(diapers[-1].weight + diapers[-1].empty_weight)
			else:
				sheet["E8"] = 0
				sheet["E9"] = 0

			for idx, urine in enumerate(night_urines):
				col = 3 + idx  # start writing from column C (3)
				try:
					time_str = urine.created_at.strftime("%H:%M") if getattr(urine, "created_at", None) is not None else ""
				except Exception:
					time_str = ""
				sheet.cell(row=15, column=col).value = time_str
				sheet.cell(row=16, column=col).value = _as_excel_number(getattr(urine, "amount", ""))
				note = (getattr(urine, "note", None) or "").strip()
				sheet.cell(row=18, column=col).value = note
				sheet.cell(row=19, column=col).value = True

			for idx, urine in enumerate(urines):
				col = 3 + idx  # start writing from column C (3)
				try:
					time_str = urine.created_at.strftime("%H:%M") if getattr(urine, "created_at", None) is not None else ""
				except Exception:
					time_str = ""
				sheet.cell(row=15, column=col).value = time_str
				sheet.cell(row=16, column=col).value = _as_excel_number(getattr(urine, "amount", ""))
				note = (getattr(urine, "note", None) or "").strip()
				sheet.cell(row=18, column=col).value = note
				sheet.cell(row=19, column=col).value = False
				
			for idx, water in enumerate(waters):
				col = 3 + idx  # start writing from column C (3)
				try:
					time_str = water.created_at.strftime("%H:%M") if getattr(water, "created_at", None) is not None else ""
				except Exception:
					time_str = ""
				sheet.cell(row=29, column=col).value = time_str
				sheet.cell(row=30, column=col).value = 200
				sheet.cell(row=31, column=col).value = "Maaltijd" if getattr(water, "meal", False) else "Drinkmoment"

			if len(waters) > (1500/200):
				sheet["J5"] = 'Ja' 

			day_urines_not_op_toilet = [
				u for u in urines if (getattr(u, "note", None) or "").strip() != "Op toilet"
			]

			if day_urines_not_op_toilet:
				sheet["J6"] = "Ja"
				sheet['J7'] = len(day_urines_not_op_toilet)
			else:
				sheet["J6"] = "Nee"
				sheet['J7'] = 0

			sheet["J8"] = len(urines)

	wb.save(diary_path)
	wb.close()

	return str(diary_path)

	